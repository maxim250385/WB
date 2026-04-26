# =============================================================================
#  modules/excel_writer.py — Сохранение в Excel (2 варианта на разных листах)
#
#  Лист 1 «По товарам» (транспонированный — РЕКОМЕНДУЕМЫЙ):
#    Строка A — Наименование (заморожено, всегда видно при прокрутке вправо)
#    Столбец 1 — Атрибуты (Категория, Бренд, Вес...)
#    Столбцы 2,3,4... — Товары
#    Удобно: копируешь один столбец = все данные одного товара
#
#  Лист 2 «Классический»:
#    Строка 1 — Атрибуты (Наименование, Категория, Бренд...)
#    Строки 2,3,4... — Товары
#    Строка 1 заморожена, столбец Наименование первый
# =============================================================================

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import PATHS, SHEET_NAME
from modules.prompts import BASE_FIELDS, CATEGORY_EXTRA_FIELDS
from modules.validator import MISSING_LABEL

# ---------------------------------------------------------------------------
# Заголовки колонок
# ---------------------------------------------------------------------------
COLUMN_HEADERS: dict[str, str] = {
    "name":                   "Наименование",
    "article_supplier":       "Артикул поставщика",
    "category":               "Категория",
    "brand":                  "Бренд",
    "country_of_origin":      "Страна производства",
    "length_cm":              "Длина (см)",
    "width_cm":               "Ширина (см)",
    "height_cm":              "Высота (см)",
    "weight_kg":              "Вес (кг)",
    "description":            "Описание (факты)",
    "complectation":          "Комплектация",
    "tnved":                  "Код ТН ВЭД",
    "barcode":                "Штрихкод",
    "warranty_months":        "Гарантия (мес.)",
    "material_composition":   "Состав ткани",
    "sizes_available":        "Размерный ряд",
    "color":                  "Цвет",
    "gender":                 "Пол",
    "season":                 "Сезон",
    "care_instructions":      "Уход",
    "style":                  "Стиль",
    "material_upper":         "Материал верха",
    "material_sole":          "Материал подошвы",
    "material_lining":        "Материал подкладки",
    "heel_height_cm":         "Высота каблука (см)",
    "power_w":                "Мощность (Вт)",
    "voltage_v":              "Напряжение (В)",
    "battery_type":           "Тип аккумулятора",
    "battery_capacity_mah":   "Ёмкость аккум. (мАч)",
    "connectivity":           "Интерфейсы",
    "model_number":           "Номер модели",
    "os":                     "ОС",
    "display_inches":         "Экран (дюймы)",
    "memory_gb":              "Память (ГБ)",
    "processor":              "Процессор",
    "camera_mp":              "Камера (МП)",
    "energy_class":           "Класс энергопотр.",
    "noise_db":               "Шум (дБ)",
    "capacity_l":             "Объём (л)",
    "spin_rpm":               "Отжим (об/мин)",
    "max_load_kg":            "Макс. загрузка (кг)",
    "wash_programs":          "Кол-во программ",
    "water_consumption_l":    "Расход воды (л)",
    "features":               "Доп. функции",
    "material_main":          "Основной материал",
    "material_finish":        "Материал отделки",
    "load_capacity_kg":       "Нагрузка макс. (кг)",
    "assembly_required":      "Требует сборки",
    "dimensions_assembled":   "Размеры в сборе",
    "style_interior":         "Стиль интерьера",
    "volume_l":               "Объём (л)",
    "volume_ml":              "Объём/вес нетто (мл/г)",
    "skin_type":              "Тип кожи/волос",
    "active_ingredients":     "Активные компоненты",
    "shelf_life_months":      "Срок годности (мес.)",
    "certification":          "Сертификация",
    "application_method":     "Способ применения",
    "effect":                 "Эффект",
    "age_from_years":         "Возраст от (лет)",
    "battery_required":       "Требует батарейки",
    "dimensions_cm":          "Размеры игрушки (ДxШxВ)",
    "developing_skills":      "Развивает навыки",
    "shelf_life_days":        "Срок годности (дней)",
    "storage_conditions":     "Условия хранения",
    "composition":            "Состав/ингредиенты",
    "calories_per_100g":      "Калорийность (100г)",
    "proteins_per_100g":      "Белки (100г)",
    "fats_per_100g":          "Жиры (100г)",
    "carbs_per_100g":         "Углеводы (100г)",
    "allergens":              "Аллергены",
    "compatible_brands":      "Совместимые марки авто",
    "compatible_models":      "Совместимые модели",
    "compatible_years":       "Годы выпуска авто",
    "installation_complexity": "Сложность установки",
    "connector_type":         "Тип разъёма",
    "waterproof":             "Водонепроницаемость",
    "temperature_min_c":      "Мин. темп. (°C)",
    "dimensions_packed":      "Размеры в сложенном виде",
    "fishing_type":           "Вид рыбалки",
    "test_g":                 "Тест приманки (г)",
    "length_m":               "Длина (м)",
    "breaking_strength_kg":   "Разрывная нагрузка (кг)",
    "gear_ratio":             "Передаточное число",
    "_source":                "Источник файла",
}

# ---------------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------------
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # тёмно-синий
_ATTR_FILL = PatternFill("solid", fgColor="2E75B6")   # синий (заголовок атрибутов)
_MISS_FILL = PatternFill("solid", fgColor="FFE0B2")   # оранжевый (уточнить)
_ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")   # голубоватый (чередование)
_NAME_FILL = PatternFill("solid", fgColor="E2EFDA")   # светло-зелёный (наименование)

_HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
_NORM_FONT = Font(size=10)
_MISS_FONT = Font(color="BF360C", italic=True, size=10)
_NAME_FONT = Font(bold=True, size=10)

_THIN = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _get_all_columns() -> list[str]:
    """Все поля: name первый, потом базовые без name, потом все категорийные."""
    seen, result = set(), []
    all_extra: list[str] = []
    for extra in CATEGORY_EXTRA_FIELDS.values():
        all_extra.extend(extra)

    # name — первый всегда
    result.append("name")
    seen.add("name")

    for f in BASE_FIELDS + all_extra + ["_source"]:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result


def _cell_style(ws, row: int, col: int,
                value, fill=None, font=None, alignment=None):
    cell            = ws.cell(row=row, column=col)
    cell.value      = value
    cell.border     = _THIN
    if fill:
        cell.fill   = fill
    if font:
        cell.font   = font
    if alignment:
        cell.alignment = alignment
    return cell


# =============================================================================
#  ЛИСТ 1 — Транспонированный «По товарам» (РЕКОМЕНДУЕМЫЙ)
#
#  Строка 1:  [пусто] | Товар 1 | Товар 2 | Товар 3 ...   ← ЗАМОРОЖЕНА
#  Строка 2:  Наименование | знач1 | знач2 | знач3 ...     ← выделена зелёным
#  Строка 3:  Артикул поставщика | ...
#  ...
#
#  При прокрутке вправо строка 1 остаётся (видно какой товар).
#  Копируешь целый столбец = все данные одного товара.
# =============================================================================

def _write_transposed(ws, records: list[dict], columns: list[str]):
    # Строка 1 — заголовок «Атрибут» + наименования товаров
    _cell_style(ws, 1, 1, "Атрибут", fill=_HDR_FILL,
                font=_HDR_FONT, alignment=_CENTER)
    for ti, record in enumerate(records, start=2):
        product_name = record.get("name") or record.get("_source") or f"Товар {ti-1}"
        _cell_style(ws, 1, ti, product_name,
                    fill=_HDR_FILL, font=_HDR_FONT, alignment=_CENTER)

    # Строки — атрибуты
    for ri, field in enumerate(columns, start=2):
        attr_label = COLUMN_HEADERS.get(field, field)

        # Стиль ячейки атрибута (колонка A)
        is_name = (field == "name")
        attr_fill = _NAME_FILL if is_name else _ATTR_FILL
        attr_font = Font(color="FFFFFF" if not is_name else "000000",
                         bold=True, size=10)
        _cell_style(ws, ri, 1, attr_label,
                    fill=attr_fill, font=attr_font, alignment=_LEFT)

        # Значения для каждого товара
        for ti, record in enumerate(records, start=2):
            value = record.get(field, "")
            alt   = _ALT_FILL if ti % 2 == 0 else PatternFill()

            if value == MISSING_LABEL:
                _cell_style(ws, ri, ti, value,
                            fill=_MISS_FILL, font=_MISS_FONT, alignment=_CENTER)
            elif is_name:
                _cell_style(ws, ri, ti, value or None,
                            fill=_NAME_FILL, font=_NAME_FONT, alignment=_LEFT)
            else:
                _cell_style(ws, ri, ti, value or None,
                            fill=alt, font=_NORM_FONT, alignment=_LEFT)

    # Ширина столбца атрибутов (A)
    max_len = max((len(COLUMN_HEADERS.get(f, f)) for f in columns), default=10)
    ws.column_dimensions["A"].width = min(max_len + 2, 40)

    # Ширина столбцов товаров
    for ti, record in enumerate(records, start=2):
        col_letter = get_column_letter(ti)
        max_w = 20
        for field in columns:
            v = str(record.get(field) or "")
            max_w = max(max_w, len(v))
        ws.column_dimensions[col_letter].width = min(max_w + 2, 50)

    # Заморозка строки 1 (шапка с названиями товаров)
    ws.freeze_panes = "B2"


# =============================================================================
#  ЛИСТ 2 — Классический (товары в строках, атрибуты в столбцах)
#
#  Строка 1:  Наименование | Артикул | Категория | ...     ← ЗАМОРОЖЕНА
#  Строка 2:  Товар 1 данные...
#  Строка 3:  Товар 2 данные...
#
#  Наименование — первый столбец, всегда видно при прокрутке вправо.
# =============================================================================

def _write_classic(ws, records: list[dict], columns: list[str]):
    # Строка 1 — заголовки атрибутов
    for ci, field in enumerate(columns, start=1):
        is_name = (field == "name")
        fill    = _NAME_FILL if is_name else _HDR_FILL
        font    = Font(color="000000" if is_name else "FFFFFF", bold=True, size=10)
        _cell_style(ws, 1, ci, COLUMN_HEADERS.get(field, field),
                    fill=fill, font=font, alignment=_CENTER)

    # Строки данных
    for ri, record in enumerate(records, start=2):
        alt = _ALT_FILL if ri % 2 == 0 else PatternFill()
        for ci, field in enumerate(columns, start=1):
            value   = record.get(field, "")
            is_name = (field == "name")
            if value == MISSING_LABEL:
                _cell_style(ws, ri, ci, value,
                            fill=_MISS_FILL, font=_MISS_FONT, alignment=_CENTER)
            elif is_name:
                _cell_style(ws, ri, ci, value or None,
                            fill=_NAME_FILL, font=_NAME_FONT, alignment=_LEFT)
            else:
                _cell_style(ws, ri, ci, value or None,
                            fill=alt, font=_NORM_FONT, alignment=_LEFT)

    # Ширина столбцов
    for ci, field in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        max_len    = len(COLUMN_HEADERS.get(field, field))
        for row in ws.iter_rows(min_col=ci, max_col=ci, min_row=2, values_only=True):
            max_len = max(max_len, len(str(row[0] or "")))
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 45))

    # Заморозка строки 1 (заголовки атрибутов)
    ws.freeze_panes = "A2"


# =============================================================================
#  Основная функция
# =============================================================================

def save_to_excel(records: list[dict], filepath: str | None = None) -> str:
    os.makedirs(PATHS["output"], exist_ok=True)
    if filepath is None:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(PATHS["output"], f"products_{ts}.xlsx")

    columns = _get_all_columns()

    wb = openpyxl.Workbook()

    # Лист 1 — транспонированный (рекомендуемый)
    ws1        = wb.active
    ws1.title  = "По товарам (рекомендуется)"
    _write_transposed(ws1, records, columns)

    # Лист 2 — классический
    ws2        = wb.create_sheet("Классический")
    _write_classic(ws2, records, columns)

    wb.save(filepath)
    return filepath
